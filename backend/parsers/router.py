import hashlib
import importlib
import logging
import mimetypes
import re
from pathlib import Path
from typing import Optional

from parsers.result import ParseResult
from schemas.uts import TransactionFlag, UniversalTransaction
from security.clamav import scan_file

logger = logging.getLogger(__name__)

# Implemented parsers — complete, tested
IMPLEMENTED_BANKS = {"sbi", "hdfc", "axis", "kotak", "indusind", "idfc", "bandhan", "yes_bank"}

BANK_KEYWORDS = {
    "sbi": ["state bank", "sbi", "sbinb"],
    "hdfc": ["hdfc", "hdfcbank"],
    "axis": ["axis", "axis bank", "axisbank", "utib"],
    "kotak": ["kotak", "kotakmahindra", "kkbk"],
    "indusind": ["induslnd", "indusind", "indus ind", "iibl", "indb"],
    "idfc":     ["idfc", "idfb", "idfc first", "idfcfirst"],
    "bandhan":  ["bandhan", "bdbl", "bandhan bank"],
    "yes_bank": ["yes bank", "yesbank", "yesb", "yes bank ltd"],
    "icici": ["icici", "icicib"],
    "pnb": ["punjab national", "pnb", "punb"],
    "canara": ["canara", "cnrb"],
    "union_bank": ["union bank", "ubin"],
}

ACCOUNT_PATTERNS = [
    re.compile(r"account\s*(?:number|no|#)?\s*[:\-]?\s*([Xx\d]{6,24})", re.IGNORECASE),
    re.compile(r"a/c\s*(?:no)?\s*[:\-]?\s*([Xx\d]{6,24})", re.IGNORECASE),
    re.compile(r"acct\s*(?:no)?\s*[:\-]?\s*([Xx\d]{6,24})", re.IGNORECASE),
]

HOLDER_PATTERNS = [
    re.compile(r"(?:account\s*holder|customer\s*name|name)\s*[:\-]?\s*([A-Z][A-Z .]{2,80})", re.IGNORECASE),
]


def detect_bank(file_path: str, first_page_text: str = "") -> Optional[str]:
    """
    Detect bank from filename and first-page text.
    Returns bank key (e.g. 'sbi') or None.
    If detection fails, the upload UI prompts the user to select manually.
    """
    stem_lower = Path(file_path).stem.lower()
    first_page_lower = first_page_text.lower()
    
    # Stage 1: Check filename first (very high confidence)
    for bank_key, keywords in BANK_KEYWORDS.items():
        if any(kw in stem_lower for kw in keywords):
            return bank_key

    # Stage 2: Parse and resolve IFSC code from the header block
    import re
    ifsc_match = re.search(r'\bifsc\s*(?:code)?\s*[:\-.\s]*\s*([a-z]{4})0[a-z0-9]{6}\b', first_page_lower)
    if ifsc_match:
        prefix = ifsc_match.group(1).upper()
        ifsc_mapping = {
            "IDFB": "idfc",
            "UTIB": "axis",
            "HDFC": "hdfc",
            "ICIC": "icici",
            "SBIN": "sbi",
            "KKBK": "kotak",
            "YESB": "yes_bank",
            "INDB": "indusind",
            "BDBL": "bandhan",
            "PUNB": "pnb",
            "CNRB": "canara",
            "UBIN": "union_bank",
        }
        if prefix in ifsc_mapping:
            return ifsc_mapping[prefix]

    # Stage 3: Look for bank keywords in the header block (first 2000 chars)
    # We prioritize matching based on owner's bank header, ignoring transaction details noise.
    header_part = first_page_lower[:2000]
    for bank_key, keywords in BANK_KEYWORDS.items():
        sorted_kws = sorted(keywords, key=len, reverse=True)
        for kw in sorted_kws:
            if kw in header_part:
                # If we matched Axis but the header also contains IDFC terms, choose IDFC
                if bank_key == "axis" and ("idfc" in header_part or "idfb" in header_part):
                    continue
                return bank_key

    # Fallback to checking the entire text if needed
    for bank_key, keywords in BANK_KEYWORDS.items():
        sorted_kws = sorted(keywords, key=len, reverse=True)
        for kw in sorted_kws:
            if kw in first_page_lower:
                if bank_key == "axis" and ("idfc" in first_page_lower or "idfb" in first_page_lower):
                    continue
                return bank_key

    return None


async def route_file(
    file_path: str,
    case_id: str,
    statement_id: str,
    bank_override: Optional[str] = None,
    original_filename: Optional[str] = None,
    progress_callback=None,
    column_mapping: Optional[dict] = None,
) -> tuple[list[UniversalTransaction], dict]:
    if progress_callback:
        await progress_callback(5, "Scanning file for malware...")

    clean, reason = await scan_file(file_path)
    if not clean:
        raise ValueError(f"File rejected by antivirus: {reason}")

    if progress_callback:
        await progress_callback(10, "Detecting bank layout...")

    file_hash = hashlib.sha256(Path(file_path).read_bytes()).hexdigest()
    ext = Path(file_path).suffix.lower()
    mime = mimetypes.guess_type(file_path)[0] or "application/octet-stream"
    first_page_text = _extract_first_page_text(file_path, ext)
    bank_key = bank_override or detect_bank(original_filename or file_path, first_page_text) or "generic"
    bank_name = bank_key.replace("_", " ").title()
    warnings: list[str] = []
    txns: list[UniversalTransaction] = []
    method = "generic"
    parser_name = "generic"

    if bank_key in IMPLEMENTED_BANKS and not column_mapping:
        try:
            if progress_callback:
                await progress_callback(20, f"Attempting specific parser for {bank_key.upper()}...")
            module = importlib.import_module(f".banks.{bank_key}", package="parsers")
            parser_name = f"bank_{bank_key}"
            method = f"bank_specific_{ext.lstrip('.') or 'file'}"
            if ext == ".pdf":
                txns = await module.parse_pdf(file_path)
            elif ext in (".xlsx", ".xls"):
                txns = await module.parse_excel(file_path)
            elif ext == ".csv":
                txns = await module.parse_csv(file_path)
            elif ext == ".docx":
                from parsers.docx_parser import parse_docx
                txns = await parse_docx(file_path, bank_key)
        except Exception as e:
            warnings.append(f"Specific parser {bank_key} failed; generic fallback used")
            logger.warning("Specific parser for %s failed, falling back to generic: %s", bank_key, e)
            txns = []

    if not txns:
        parser_name = "generic"
        method = f"generic_{ext.lstrip('.') or 'file'}"
        if progress_callback:
            await progress_callback(25, f"Routing to generic {ext.upper()} parsing pipeline...")
        try:
            if ext == ".pdf":
                txns = await _generic_parse_pdf(file_path, bank_name, file_hash, progress_callback, column_mapping, statement_id)
                method = "ocr_pdf" if any(getattr(t, "ocr_confidence", None) is not None for t in txns) else "digital_pdf"
            elif ext in (".xlsx", ".xls"):
                if progress_callback:
                    await progress_callback(30, "Parsing Excel sheets and resolving merged cells...")
                txns = await _generic_parse_excel(file_path, bank_name, file_hash, column_mapping, statement_id)
            elif ext == ".csv":
                if progress_callback:
                    await progress_callback(30, "Sniffing CSV delimiters and parsing table...")
                txns = await _generic_parse_csv(file_path, bank_name, file_hash, column_mapping, statement_id)
            elif ext == ".docx":
                if progress_callback:
                    await progress_callback(30, "Parsing Word document tables...")
                txns = await _generic_parse_docx(file_path, bank_name, file_hash, column_mapping, statement_id)
            elif ext in (".png", ".jpg", ".jpeg", ".tiff", ".webp", ".bmp"):
                if progress_callback:
                    await progress_callback(40, "Running OCR on image file...")
                from parsers.pdf_scanned import parse_image_file
                txns = await parse_image_file(file_path, bank_name, column_mapping=column_mapping)
                method = "ocr_image"
            else:
                raise ValueError(f"Unsupported file type: {ext}")
        except Exception as e:
            logger.error("Generic parser pipeline failed: %s", e)
            raise ValueError(f"Failed to parse statement: {e}")

    account_id, account_holder = _extract_account_metadata(first_page_text)
    txns, normalize_warnings = _normalize_transactions(
        txns,
        case_id=case_id,
        statement_id=statement_id,
        file_hash=file_hash,
        bank_name=bank_name,
        parser_name=parser_name,
        method=method,
        fallback_account_id=account_id,
        fallback_account_holder=account_holder,
    )
    warnings.extend(normalize_warnings)

    ocr_used = any(getattr(t, "ocr_confidence", None) is not None for t in txns)
    result = ParseResult.from_transactions(
        txns,
        parser_name=parser_name,
        method=method,
        ocr_used=ocr_used,
        warnings=warnings,
        detected_mapping=column_mapping,
    )
    metadata = result.metadata(bank_name=bank_name, file_hash=file_hash, mime_type=mime)
    return result.transactions, metadata


def _extract_first_page_text(file_path: str, ext: str) -> str:
    try:
        if ext == ".pdf":
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                return (pdf.pages[0].extract_text() or "") if pdf.pages else ""
    except Exception:
        pass
    return ""


def _extract_account_metadata(text: str) -> tuple[str, str]:
    account_id = ""
    account_holder = ""
    for pattern in ACCOUNT_PATTERNS:
        m = pattern.search(text or "")
        if m:
            account_id = m.group(1).strip()
            break
    for pattern in HOLDER_PATTERNS:
        m = pattern.search(text or "")
        if m:
            account_holder = re.sub(r"\s+", " ", m.group(1)).strip()
            break
    return account_id, account_holder


def _normalize_transactions(
    txns: list[UniversalTransaction],
    case_id: str,
    statement_id: str,
    file_hash: str,
    bank_name: str,
    parser_name: str,
    method: str,
    fallback_account_id: str = "",
    fallback_account_holder: str = "",
) -> tuple[list[UniversalTransaction], list[str]]:
    warnings: list[str] = []
    normalized: list[UniversalTransaction] = []
    identity_fallback_used = False

    for idx, txn in enumerate(txns or []):
        txn.case_id = case_id
        txn.statement_id = statement_id
        txn.source_file_hash = file_hash
        txn.bank_name = txn.bank_name or bank_name
        txn.account_holder = txn.account_holder or fallback_account_holder or ""
        txn.parser_name = txn.parser_name or parser_name
        txn.raw_row_index = txn.raw_row_index if txn.raw_row_index is not None else idx
        if txn.raw_row_json is None:
            txn.raw_row_json = {
                "row_index": txn.raw_row_index,
                "date": txn.txn_date.isoformat() if txn.txn_date else None,
                "amount": str(txn.amount),
                "txn_type": str(txn.txn_type),
                "balance_after": str(txn.balance_after) if txn.balance_after is not None else None,
                "narration": txn.narration,
            }
        if not txn.account_id or not txn.account_id.strip():
            txn.account_id = fallback_account_id or f"STATEMENT-{statement_id[:8]}"
            txn.identity_confidence = 0.75 if fallback_account_id else 0.20
            identity_fallback_used = True
        elif txn.identity_confidence is None:
            txn.identity_confidence = 0.90
        if txn.parse_confidence is None:
            txn.parse_confidence = txn.ocr_confidence if txn.ocr_confidence is not None else 0.90

        seed = "|".join([
            statement_id,
            str(txn.raw_row_index),
            txn.account_id or "",
            txn.txn_date.isoformat() if txn.txn_date else "",
            str(txn.amount),
            str(txn.txn_type),
            str(txn.balance_after) if txn.balance_after is not None else "",
            txn.narration or "",
            parser_name,
        ])
        txn.txn_hash = hashlib.sha256(seed.encode()).hexdigest()
        normalized.append(txn)

    if identity_fallback_used:
        warnings.append("Account identity could not be fully extracted; statement-scoped fallback account id used")
    return normalized, warnings


async def _generic_parse_excel(file_path: str, bank_name: str, file_hash: str, column_mapping: Optional[dict] = None, statement_id: str = "") -> list[UniversalTransaction]:
    ext = Path(file_path).suffix.lower()
    sheets_data = {}
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            rows = []
            for rx in range(sheet.nrows):
                rows.append([str(sheet.cell_value(rx, cx)).strip() for cx in range(sheet.ncols)])
            sheets_data[sheet.name] = rows
    else:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet in wb.worksheets:
            rows = []
            merged_val_map = {}
            for rng in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = rng.bounds
                top_left_cell_val = sheet.cell(row=min_row, column=min_col).value
                for r_idx in range(min_row, max_row + 1):
                    for c_idx in range(min_col, max_col + 1):
                        merged_val_map[(r_idx, c_idx)] = top_left_cell_val
            for r_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                row_cells = []
                for c_idx, cell in enumerate(row, start=1):
                    val = merged_val_map.get((r_idx, c_idx), cell.value)
                    row_cells.append(str(val or "").strip())
                rows.append(row_cells)
            sheets_data[sheet.title] = rows

    from parsers.generic_parser import parse_date, parse_decimal, parse_generic_table
    best_sheet_name = None
    best_sheet_score = -1
    for sheet_name, rows in sheets_data.items():
        score = 0
        for row in rows[:100]:
            for cell in row:
                if parse_date(cell):
                    score += 2
                if parse_decimal(cell) is not None:
                    score += 1
        if score > best_sheet_score:
            best_sheet_score = score
            best_sheet_name = sheet_name
    best_rows = sheets_data[best_sheet_name] if best_sheet_name else []
    return parse_generic_table(best_rows, bank_name, file_hash, column_mapping, statement_id=statement_id)


async def _generic_parse_csv(file_path: str, bank_name: str, file_hash: str, column_mapping: Optional[dict] = None, statement_id: str = "") -> list[UniversalTransaction]:
    import csv
    import chardet

    with open(file_path, "rb") as f:
        sample_bytes = f.read(8192)
    enc = chardet.detect(sample_bytes)["encoding"] or "utf-8"
    try:
        sample_text = sample_bytes.decode(enc, errors="replace")
    except Exception:
        sample_text = sample_bytes.decode("utf-8", errors="replace")
        enc = "utf-8"

    delimiters = [",", ";", "\t", "|"]
    delimiter_counts = {d: 0 for d in delimiters}
    for line in sample_text.split("\n")[:10]:
        for delimiter in delimiters:
            delimiter_counts[delimiter] += line.count(delimiter)
    chosen_delimiter = max(delimiter_counts, key=delimiter_counts.get)
    if delimiter_counts[chosen_delimiter] == 0:
        chosen_delimiter = ","

    rows = []
    with open(file_path, encoding=enc, errors="replace") as f:
        reader = csv.reader(f, delimiter=chosen_delimiter)
        for row in reader:
            if any(row):
                rows.append([str(c or "").strip() for c in row])

    from parsers.generic_parser import parse_generic_table
    return parse_generic_table(rows, bank_name, file_hash, column_mapping, statement_id=statement_id, pre_cleaned=True)


async def _generic_parse_docx(file_path: str, bank_name: str, file_hash: str, column_mapping: Optional[dict] = None, statement_id: str = "") -> list[UniversalTransaction]:
    from docx import Document
    from parsers.generic_parser import parse_generic_table

    doc = Document(file_path)
    rows = []
    for table in doc.tables:
        for row in table.rows:
            row_cells = [cell.text.strip() for cell in row.cells]
            if any(row_cells):
                rows.append(row_cells)
    return parse_generic_table(rows, bank_name, file_hash, column_mapping, statement_id=statement_id)


async def _generic_parse_pdf(file_path: str, bank_name: str, file_hash: str, progress_callback=None, column_mapping: Optional[dict] = None, statement_id: str = "") -> list[UniversalTransaction]:
    from parsers.generic_parser import parse_generic_table
    from parsers.pdf_scanned import is_pdf_scanned, parse_scanned_pdf
    import re

    if is_pdf_scanned(file_path):
        if progress_callback:
            await progress_callback(80, "PDF is scanned. Running OCR fallback...")
        return await parse_scanned_pdf(file_path, bank_name, progress_callback, column_mapping=column_mapping)

    extraction_attempts = []
    try:
        if progress_callback:
            await progress_callback(30, "Stage 1/5: Extracting PDF tables (pdfplumber)...")
        import pdfplumber
        rows = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    for row in table:
                        cells = [str(c or "").strip() for c in row]
                        if any(cells):
                            rows.append(cells)
        extraction_attempts.append(("pdfplumber_table", rows))
    except Exception as e:
        logger.debug("pdfplumber table extract failed: %s", e)

    for flavor, pct, label in (("lattice", 45, "bordered"), ("stream", 60, "borderless")):
        try:
            if progress_callback:
                await progress_callback(pct, f"Stage {2 if flavor == 'lattice' else 3}/5: Extracting {label} tables (camelot {flavor})...")
            import camelot
            rows = []
            tables = camelot.read_pdf(file_path, pages="all", flavor=flavor)
            for table in tables:
                for _, row in table.df.iterrows():
                    cells = [str(c or "").strip() for c in row]
                    if any(cells):
                        rows.append(cells)
            extraction_attempts.append((f"camelot_{flavor}", rows))
        except Exception as e:
            logger.debug("camelot %s extract failed: %s", flavor, e)

    try:
        if progress_callback:
            await progress_callback(75, "Stage 4/5: Parsing PDF text line-by-line...")
        import pdfplumber
        rows = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                for line in text.split("\n"):
                    tokens = re.split(r"\s{2,}", line.strip())
                    if len(tokens) <= 1:
                        tokens = line.strip().split("\t")
                    tokens = [t.strip() for t in tokens if t.strip()]
                    if len(tokens) >= 2:
                        rows.append(tokens)
        extraction_attempts.append(("pdf_text", rows))
    except Exception as e:
        logger.debug("pdfplumber line-by-line text parsing failed: %s", e)

    for attempt_name, rows in extraction_attempts:
        if not rows:
            continue
        txns = parse_generic_table(rows, bank_name, file_hash, column_mapping, statement_id=statement_id)
        if txns:
            for txn in txns:
                txn.parser_name = attempt_name
            logger.info("Successfully parsed PDF using %s", attempt_name)
            return txns

    if progress_callback:
        await progress_callback(85, "Stage 5/5: Running OCR scanned document fallback...")
    return await parse_scanned_pdf(file_path, bank_name, progress_callback, column_mapping=column_mapping)