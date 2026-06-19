"""
Comprehensive test script for robust parsing features.
Tests:
1. Indian/European number formats and parenthesized negative values.
2. Multiple date formats.
3. CSV delimiter and encoding sniffer.
4. Excel merged cell resolution and sheet selection density scoring.
5. Multi-line narration merging.
6. Mathematical column classification via running balances.
"""
import unittest
import tempfile
import csv
from pathlib import Path
from decimal import Decimal
from datetime import datetime
import openpyxl

from parsers.generic_parser import (
    clean_amount_str,
    parse_decimal,
    parse_date,
    parse_generic_table,
    infer_column_roles
)
from parsers.router import _generic_parse_csv, _generic_parse_excel

class TestRobustParsers(unittest.TestCase):

    def test_amount_parsing(self):
        # Indian formats
        self.assertEqual(parse_decimal("₹ 1,23,456.78"), Decimal("123456.78"))
        self.assertEqual(parse_decimal("12,34,567.89"), Decimal("1234567.89"))
        self.assertEqual(parse_decimal(" Rs. 15,000.00 Cr "), Decimal("15000.00"))
        
        # Parentheses (negative)
        self.assertEqual(parse_decimal("(1,234.56)"), Decimal("-1234.56"))
        self.assertEqual(parse_decimal(" (₹ 15,200.50) "), Decimal("-15200.50"))
        
        # Standard cleaning
        self.assertEqual(parse_decimal("- 100.50"), Decimal("-100.50"))
        self.assertEqual(parse_decimal("1.234.567,89"), Decimal("1234567.89"))  # European format

    def test_date_parsing(self):
        self.assertEqual(parse_date("15/06/2026"), datetime(2026, 6, 15))
        self.assertEqual(parse_date("15-06-2026"), datetime(2026, 6, 15))
        self.assertEqual(parse_date("15.06.26"), datetime(2026, 6, 15))
        self.assertEqual(parse_date("15 Jun 2026"), datetime(2026, 6, 15))
        self.assertEqual(parse_date("15-June-2026"), datetime(2026, 6, 15))

    def test_csv_sniffer(self):
        # Create temp CSV with pipe delimiter
        with tempfile.NamedTemporaryFile(suffix=".csv", mode="w", newline="", delete=False) as tmp:
            writer = csv.writer(tmp, delimiter="|")
            writer.writerow(["Date", "Description", "Amount", "Balance"])
            writer.writerow(["10/06/2026", "Salary", "50000.00", "50000.00"])
            writer.writerow(["11/06/2026", "Rent", "-15000.00", "35000.00"])
            tmp_path = tmp.name

        try:
            import asyncio
            txns = asyncio.run(_generic_parse_csv(tmp_path, "TestBank", "hash"))
            self.assertEqual(len(txns), 2)
            self.assertEqual(txns[0].amount, Decimal("50000.00"))
            self.assertEqual(txns[1].amount, Decimal("15000.00"))
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    def test_excel_merged_cells_and_sheet_selection(self):
        wb = openpyxl.Workbook()
        # Sheet 1: Dummy sheet with low density
        ws1 = wb.active
        ws1.title = "DummySheet"
        ws1["A1"] = "Nothing here"
        
        # Sheet 2: Main transaction sheet with merged cells
        ws2 = wb.create_sheet(title="Transactions")
        ws2["A1"] = "Statement Date"
        ws2["B1"] = "Description"
        ws2["C1"] = "Debit"
        ws2["D1"] = "Credit"
        ws2["E1"] = "Balance"
        
        ws2["A2"] = "12/06/2026"
        ws2["B2"] = "UPI Transfer to John"
        ws2["C2"] = "1000.00"
        ws2["E2"] = "9000.00"
        
        # Merged range for A3:A4
        ws2.merge_cells("A3:A4")
        ws2["A3"] = "13/06/2026"
        
        # Merged cells need to propagate date
        ws2["B3"] = "Rent payment"
        ws2["C3"] = "5000.00"
        ws2["E3"] = "4000.00"
        
        ws2["B4"] = "Interest Received"
        ws2["D4"] = "50.00"
        ws2["E4"] = "4050.00"
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
            
        try:
            import asyncio
            txns = asyncio.run(_generic_parse_excel(tmp_path, "TestBank", "hash"))
            self.assertEqual(len(txns), 3)
            # Check sheet selection (should pick Transactions sheet)
            # Check merged cells value propagation (both index 1 and 2 should have date 13/06/2026)
            self.assertEqual(txns[0].txn_date, datetime(2026, 6, 12))
            self.assertEqual(txns[1].txn_date, datetime(2026, 6, 13))
            self.assertEqual(txns[2].txn_date, datetime(2026, 6, 13))
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    def test_multiline_narration(self):
        rows = [
            ["Date", "Description", "Amount", "Balance"],
            ["14/06/2026", "UPI Transfer to John", "-1000", "9000"],
            ["", "Ref ID: 1234567890", "", ""],
            ["", "Remarks: For Dinner", "", ""],
            ["15/06/2026", "Salary Credit", "50000", "59000"]
        ]
        txns = parse_generic_table(rows, "TestBank", "hash")
        self.assertEqual(len(txns), 2)
        self.assertEqual(txns[0].narration, "UPI Transfer to John Ref ID: 1234567890 Remarks: For Dinner")
        self.assertEqual(txns[1].narration, "Salary Credit")

    def test_mathematical_column_classification(self):
        # No headers, raw table with Date, Amount1, Amount2, Balance
        # Let's see if mathematical deltas identify C1=Debit, C2=Credit
        rows = [
            ["14/06/2026", "UPI Transfer", "1000", "", "9000"],
            ["15/06/2026", "Salary", "", "50000", "59000"],
            ["16/06/2026", "Coffee", "150", "", "58850"]
        ]
        mapping = infer_column_roles(rows)
        # c0 = date
        # c1 = 1000, None, 150 (Debit)
        # c2 = None, 50000, None (Credit)
        # c3 = balance
        # c4 = narration
        self.assertEqual(mapping.get("date"), 0)
        self.assertEqual(mapping.get("debit"), 2)
        self.assertEqual(mapping.get("credit"), 3)
        self.assertEqual(mapping.get("balance"), 4)

if __name__ == "__main__":
    unittest.main()
